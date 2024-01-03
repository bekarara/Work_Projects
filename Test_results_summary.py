import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter.filedialog import askopenfilename
import tkinter as tk


# select input xlsx file:
input_filepath = askopenfilename()

#populate summary sheet with list of sheet names
xl = pd.ExcelFile(input_filepath)
sheet_names = xl.sheet_names


# read input file excel info as dataframe
excel_data_df_list = []
for sheets in sheet_names:
    df = pd.read_excel(input_filepath, sheet_name = sheets)
    excel_data_df_list.append(df)

#initialize result dataframe
summary_df = pd.DataFrame(columns = ['Test Case ID','Pass','Fail','Open','Pass /w comments','Partial'])


#count the number of pass and fail results
for index, test_results in enumerate(excel_data_df_list):
    num_pass = (test_results[4] == "Pass").sum() #test results is in column "4"
    num_fail = (test_results[4] == "Fail").sum()
    num_open = (test_results[4] == "Open").sum()
    num_comments = (test_results[4] == "Pass /w comments").sum()
    num_partial = (test_results[4] == "Partial").sum()
    new_df_row = [sheet_names[index],num_pass,num_fail,num_open,num_comments,num_partial]
    summary_df.loc[len(summary_df)] = new_df_row  

#add summary sheet
wb = load_workbook(input_filepath)
summary_sheet = wb.create_sheet("Summary",0)
wb.save(input_filepath)
wb.close()

with pd.ExcelWriter(input_filepath, engine = 'openpyxl', mode = 'a', if_sheet_exists='replace') as writer:
    summary_df.to_excel(writer,sheet_name = 'Summary', index = None)
